from datetime import date, timedelta
import calendar
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from .models import Teacher, TimetableCard, LessonRecord, MonthlyReport

PERIOD_TIMES = {
    "1": ("09:00", "10:10"),
    "2": ("10:20", "11:30"),
    "3": ("11:40", "12:50"),
    "4": ("13:30", "14:40"),
    "5": ("14:50", "16:00"),
    "6": ("16:10", "17:20"),
    "7": ("17:30", "18:40"),
    "8": ("18:50", "20:00"),
}

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
PERIODS = ["1", "2", "3", "4", "5", "6", "7", "8"]


def get_lesson_type(classroom_name):
    if classroom_name and "lecture" in classroom_name.lower():
        return "lecture"
    return "seminar"


def login_view(request):
    error = None
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            teacher = Teacher.objects.filter(user=user).first()
            if teacher and teacher.is_dean:
                return redirect("dean_dashboard")
            return redirect("timetable")
        else:
            error = "Invalid username or password"
    return render(request, "login.html", {"error": error})


def logout_view(request):
    logout(request)
    return redirect("login")

from django.views.decorators.csrf import ensure_csrf_cookie

@ensure_csrf_cookie
@login_required(login_url="login")
def timetable_view(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher:
        return render(request, "timetable.html", {"error": "No timetable linked to your account"})

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    num_days = calendar.monthrange(year, month)[1]
    month_days = [date(year, month, d) for d in range(1, num_days + 1)]
    month_days = [d for d in month_days if d.weekday() < 6]

    # Group into weeks
    weeks = []
    current_week = []
    for d in month_days:
        if d.weekday() == 0 and current_week:
            weeks.append(current_week)
            current_week = []
        current_week.append(d)
    if current_week:
        weeks.append(current_week)

    cards = TimetableCard.objects.filter(teacher=teacher).prefetch_related("subject", "classroom")

    period_list = [
        {"number": p, "start": PERIOD_TIMES[p][0], "end": PERIOD_TIMES[p][1]}
        for p in PERIODS
    ]

    # Get existing lesson records for this month
    records = LessonRecord.objects.filter(
        teacher=teacher,
        date__year=year,
        date__month=month
    )
    record_map = {(r.card_id, str(r.date)): r for r in records}

    can_submit = True  # always can submit

    # Check if already submitted
    report = MonthlyReport.objects.filter(teacher=teacher, year=year, month=month).first()
    already_submitted = report and report.status == "submitted"

    weeks_data = []
    for week in weeks:
        rows = []
        for d in week:
            day_name = DAYS[d.weekday()]
            cells = []
            for period in PERIODS:
                card = cards.filter(period=period, day=day_name).first()
                record = record_map.get((card.id if card else None, str(d)))

                # Freeze past days — create snapshot so EduPage changes don't affect them
                if card and d < today and not record:
                    record, _ = LessonRecord.objects.get_or_create(
                        card=card,
                        date=d,
                        defaults={
                            "teacher": teacher,
                            "is_covered": None,
                            "is_replaced": False,
                        }
                    )
                    record_map[(card.id, str(d))] = record

                cells.append({
                    "card": card if (d >= today or record) else None,
                    "record": record,
                    "date": str(d),
                    "period": period,
                })
            rows.append({
                "date": d,
                "day_name": day_name,
                "cells": cells,
            })
        weeks_data.append(rows)

    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month = date(year, month, num_days) + timedelta(days=1)

    return render(request, "timetable.html", {
        "teacher": teacher,
        "period_list": period_list,
        "weeks_data": weeks_data,
        "month_name": date(year, month, 1).strftime("%B %Y"),
        "year": year,
        "month": month,
        "prev": {"year": prev_month.year, "month": prev_month.month},
        "next": {"year": next_month.year, "month": next_month.month},
        "can_submit": can_submit,
        "already_submitted": already_submitted,

    })


@login_required(login_url="login")
@require_POST
def mark_lesson(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    card_id = request.POST.get("card_id")
    date_str = request.POST.get("date")
    is_covered = request.POST.get("is_covered")  # "true" or "false"
    note = request.POST.get("note", "")
    replacement_date = request.POST.get("replacement_date", "")

    card = get_object_or_404(TimetableCard, id=card_id, teacher=teacher)
    lesson_date = date.fromisoformat(date_str)

    record, _ = LessonRecord.objects.update_or_create(
        card=card,
        date=lesson_date,
        defaults={
            "teacher": teacher,
            "is_covered": is_covered == "true",
            "note": note,
            "is_replaced": bool(replacement_date),
            "replacement_date": date.fromisoformat(replacement_date) if replacement_date else None,
        }
    )

    return JsonResponse({"status": "ok", "is_covered": record.is_covered})


@login_required(login_url="login")
@require_POST
def submit_report(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    year = int(request.POST.get("year"))
    month = int(request.POST.get("month"))

    report, _ = MonthlyReport.objects.update_or_create(
        teacher=teacher,
        year=year,
        month=month,
        defaults={
            "status": "submitted",
            "submitted_at": timezone.now(),
        }
    )

    return JsonResponse({"status": "submitted"})


@login_required(login_url="login")
def dean_dashboard(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher or not teacher.is_dean:
        return redirect("timetable")

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    num_days = calendar.monthrange(year, month)[1]

    # Get all days in month
    all_days = [date(year, month, d) for d in range(1, num_days + 1)]
    working_days = [d for d in all_days if d.weekday() < 6]


    # Get teachers in same department
    dept_teachers = Teacher.objects.filter(
        department=teacher.department,
        user__isnull=False
    ).order_by("full_name")

    # Build grid data
    teacher_rows = []
    for t in dept_teachers:
        # Get all lesson records for this month
        records = LessonRecord.objects.filter(
            teacher=t,
            date__year=year,
            date__month=month,
            is_covered=True
        )

        # Count covered lessons per day
        day_counts = {}
        for d in working_days:
            count = records.filter(date=d).count()
            day_counts[str(d)] = count

        total = records.count()
        report = MonthlyReport.objects.filter(teacher=t, year=year, month=month).first()

        teacher_rows.append({
            "teacher": t,
            "day_counts": day_counts,
            "total": total,
            "status": report.status if report else "not submitted",
        })

    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month_date = date(year, month, num_days) + timedelta(days=1)

    return render(request, "dean_dashboard.html", {
        "teacher": teacher,
        "teacher_rows": teacher_rows,
        "working_days": working_days,
        "month_name": date(year, month, 1).strftime("%B %Y"),
        "year": year,
        "month": month,

        "prev": {"year": prev_month.year, "month": prev_month.month},
        "next": {"year": next_month_date.year, "month": next_month_date.month},
    })



import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@login_required(login_url="login")
def export_dean_excel(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher or not teacher.is_dean:
        return redirect("timetable")

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    num_days = calendar.monthrange(year, month)[1]
    all_days = [date(year, month, d) for d in range(1, num_days + 1)]
    working_days = [d for d in all_days if d.weekday() < 6]

    dept_teachers = Teacher.objects.filter(
        department=teacher.department,
        user__isnull=False
    ).order_by("full_name")

    # Build data
    teacher_rows = []
    for t in dept_teachers:
        records = LessonRecord.objects.filter(
            teacher=t,
            date__year=year,
            date__month=month,
            is_covered=True
        )
        day_counts = {str(d): records.filter(date=d).count() for d in working_days}
        total = records.count()
        teacher_rows.append({
            "name": t.full_name,
            "day_counts": day_counts,
            "total": total,
        })

    # Build Excel
    wb = Workbook()
    ws = wb.active
    month_name = date(year, month, 1).strftime("%B %Y")
    dept_name = teacher.department.name if teacher.department else "Department"
    ws.title = month_name

    # Styles
    header_fill = PatternFill("solid", start_color="1B497D", end_color="1B497D")
    total_fill = PatternFill("solid", start_color="163D6A", end_color="163D6A")
    has_lesson_fill = PatternFill("solid", start_color="E8F5EC", end_color="E8F5EC")
    white_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    blue_bold_font = Font(name="Arial", bold=True, color="1B497D", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    last_col_letter = get_column_letter(len(working_days) + 2)
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = f"{dept_name} — {month_name}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="1B497D")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # Header row
    ws.row_dimensions[2].height = 36
    ws["A2"].value = "Teacher"
    ws["A2"].font = white_font
    ws["A2"].fill = header_fill
    ws["A2"].alignment = left
    ws["A2"].border = border
    ws.column_dimensions["A"].width = 28

    for i, d in enumerate(working_days):
        col = i + 2
        col_letter = get_column_letter(col)
        cell = ws.cell(row=2, column=col)
        cell.value = f"{d.day}\n{d.strftime('%a')}"
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[col_letter].width = 6

    # Total column
    total_col = len(working_days) + 2
    total_col_letter = get_column_letter(total_col)
    total_cell = ws.cell(row=2, column=total_col)
    total_cell.value = "Total"
    total_cell.font = white_font
    total_cell.fill = total_fill
    total_cell.alignment = center
    total_cell.border = border
    ws.column_dimensions[total_col_letter].width = 8

    # Data rows
    for row_idx, row in enumerate(teacher_rows):
        excel_row = row_idx + 3
        ws.row_dimensions[excel_row].height = 22

        # Teacher name
        name_cell = ws.cell(row=excel_row, column=1)
        name_cell.value = row["name"]
        name_cell.font = bold_font
        name_cell.alignment = left
        name_cell.border = border
        name_cell.fill = PatternFill("solid", start_color="F7F9FC", end_color="F7F9FC")

        # Day counts
        for i, d in enumerate(working_days):
            col = i + 2
            count = row["day_counts"].get(str(d), 0)
            cell = ws.cell(row=excel_row, column=col)
            cell.value = count if count else None
            cell.font = bold_font if count else normal_font
            cell.alignment = center
            cell.border = border
            if count:
                cell.fill = has_lesson_fill

        # Total
        total_cell = ws.cell(row=excel_row, column=total_col)
        total_cell.value = row["total"]
        total_cell.font = blue_bold_font
        total_cell.alignment = center
        total_cell.border = border
        total_cell.fill = PatternFill("solid", start_color="EEF4FF", end_color="EEF4FF")

    # Save to response
    filename = f"{dept_name}_{month_name}.xlsx".replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.read())
    return response